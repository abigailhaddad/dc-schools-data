#!/usr/bin/env python3
"""
profile_files.py — pull every data file and inventory its structure.

Reads data_files.yaml, downloads each file (cached under data_cache/), and
records what's actually inside:

  * xlsx / xls : every sheet/tab name, the column headers of each sheet, row count
  * csv        : column headers, row count
  * pdf        : page count + extracted text/table preview from the first pages

Results are written to file_profiles.json, which generate_readme.py folds into
the catalog so each data file shows its tabs/columns inline.

Usage:
    python profile_files.py                # download missing + profile all
    python profile_files.py --refresh      # re-download everything
    python profile_files.py --only osse-finance-2024 opendata-dcps-locations-csv
    python profile_files.py --no-download   # profile only what's already cached

Files that are NOT direct downloads (Box / Egnyte / JS-rendered "about" pages)
can't be fetched with plain HTTP — they're recorded with status "needs_browser"
so the gap is visible rather than silently skipped. Harvest those with the
update-dc-schools-data skill (Playwright).
"""
from __future__ import annotations

import argparse
import io
import json
import sys
from pathlib import Path

import yaml

try:
    import requests
except ImportError:
    sys.exit("requests is required: pip install requests")

ROOT = Path(__file__).resolve().parent
DATA_FILES = ROOT / "data_files.yaml"
CACHE = ROOT / "data_cache"
OUT = ROOT / "file_profiles.json"

# URLs that don't serve a file directly over plain HTTP and must be resolved in a
# browser first (see the update-dc-schools-data skill). A Box *share page*
# (app.box.com/s/...) needs a browser to discover its file_id; once resolved to a
# box_download_shared_file URL it downloads fine with plain requests, so that form
# is NOT flagged.
NEEDS_BROWSER_PATTERNS = ("box.com/s/", "egnyte.com", "docs.google.com", "/folder/")

MAX_COLS = 60             # cap columns recorded per sheet (some files are very wide)
PDF_PREVIEW_PAGES = 3     # how many pages of a PDF to grab a text preview from
PDF_MAX_TABLE_PAGES = 120 # scan up to this many pages for tables (runtime bound)
PDF_MAX_TABLES = 25       # store at most this many sample tables (count is exact)
# Box and some gov hosts reject non-browser agents, so present as a browser.
HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                         "AppleWebKit/537.36 (KHTML, like Gecko) "
                         "Chrome/124.0 Safari/537.36 dc-schools-data-catalog/1.0"}


# --------------------------------------------------------------------------- #
# download
# --------------------------------------------------------------------------- #
def local_path(entry: dict) -> Path:
    ext = entry["kind"]
    return CACHE / f"{entry['id']}.{ext}"

def needs_browser(url: str) -> bool:
    if "box_download_shared_file" in url:   # already-resolved Box download URL
        return False
    return any(p in url for p in NEEDS_BROWSER_PATTERNS)

def download(entry: dict, refresh: bool) -> tuple[Path | None, str | None]:
    """Returns (path, error). path is None when the file couldn't be fetched."""
    url = entry["url"]
    if needs_browser(url):
        return None, "needs_browser"
    dest = local_path(entry)
    if dest.exists() and not refresh:
        return dest, None
    CACHE.mkdir(exist_ok=True)
    try:
        resp = requests.get(url, headers=HEADERS, timeout=90, allow_redirects=True)
        resp.raise_for_status()
    except Exception as exc:  # noqa: BLE001 - we want to record any failure
        return None, f"download_failed: {exc}"
    # A returned HTML page usually means a redirect to a landing/login page.
    ctype = resp.headers.get("content-type", "")
    if entry["kind"] != "csv" and "text/html" in ctype:
        return None, f"got_html_not_file (content-type={ctype})"
    dest.write_bytes(resp.content)
    return dest, None


# --------------------------------------------------------------------------- #
# profilers
# --------------------------------------------------------------------------- #
def _clean(values) -> list[str]:
    out = []
    for v in values:
        if v is None:
            continue
        s = " ".join(str(v).split())   # collapse internal newlines/whitespace
        if s:
            out.append(s)
    return out

def profile_xlsx(path: Path) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header = []
        for row in rows:                       # first non-empty row = header
            cells = _clean(row)
            if cells:
                header = cells
                break
        n_rows = ws.max_row or 0
        sheets.append({
            "name": ws.title,
            "n_rows": n_rows,
            "n_columns": len(header),
            "columns": header[:MAX_COLS],
            "columns_truncated": len(header) > MAX_COLS,
        })
    wb.close()
    return {"kind": "xlsx", "n_sheets": len(sheets), "sheets": sheets}

def profile_csv(path: Path) -> dict:
    import pandas as pd
    # nrows=0 reads just the header cheaply; then count rows separately.
    head = pd.read_csv(path, nrows=0, dtype=str, encoding_errors="replace")
    cols = list(head.columns)
    with path.open("r", errors="replace") as fh:
        n_rows = max(sum(1 for _ in fh) - 1, 0)
    return {
        "kind": "csv",
        "n_rows": n_rows,
        "n_columns": len(cols),
        "columns": cols[:MAX_COLS],
        "columns_truncated": len(cols) > MAX_COLS,
    }

def profile_pdf(path: Path) -> dict:
    """Page count + text preview (first pages) + a scan of EVERY page for tables.

    DC reports often bury the real data (by-school attendance, discipline counts)
    in appendix tables deep in the PDF, so we scan the whole document — capped at
    PDF_MAX_TABLE_PAGES for runtime — rather than just the first few pages.
    """
    from pypdf import PdfReader
    reader = PdfReader(str(path))
    n_pages = len(reader.pages)
    result = {"kind": "pdf", "n_pages": n_pages}

    text_preview, tables = [], []
    pages_with_tables = set()
    try:
        import pdfplumber
        with pdfplumber.open(str(path)) as pdf:
            for i, page in enumerate(pdf.pages):
                if i < PDF_PREVIEW_PAGES:
                    txt = (page.extract_text() or "").strip()
                    if txt:
                        text_preview.append(txt[:1500])
                if i >= PDF_MAX_TABLE_PAGES:
                    continue
                for tbl in page.extract_tables() or []:
                    if tbl and tbl[0] and len(tbl) > 1:
                        pages_with_tables.add(page.page_number)
                        if len(tables) < PDF_MAX_TABLES:   # store a sample, count all
                            tables.append({
                                "page": page.page_number,
                                "header": _clean(tbl[0])[:MAX_COLS],
                                "n_rows": len(tbl),
                            })
    except Exception:
        for page in reader.pages[:PDF_PREVIEW_PAGES]:
            txt = (page.extract_text() or "").strip()
            if txt:
                text_preview.append(txt[:1500])

    result["text_preview"] = text_preview
    result["tables"] = tables
    result["n_table_pages"] = len(pages_with_tables)
    result["table_scan_capped"] = n_pages > PDF_MAX_TABLE_PAGES
    # honest signal: does this PDF actually carry extractable tabular data?
    result["has_tabular_data"] = len(pages_with_tables) > 0
    return result

PROFILERS = {
    "xlsx": profile_xlsx,
    "xls": profile_xlsx,
    "csv": profile_csv,
    "pdf": profile_pdf,
}


# --------------------------------------------------------------------------- #
# main
# --------------------------------------------------------------------------- #
def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--refresh", action="store_true", help="re-download cached files")
    ap.add_argument("--no-download", action="store_true", help="profile only cached files")
    ap.add_argument("--only", nargs="*", help="profile only these file ids")
    args = ap.parse_args()

    entries = yaml.safe_load(DATA_FILES.read_text())["data_files"]
    if args.only:
        entries = [e for e in entries if e["id"] in set(args.only)]

    # Merge into any existing profiles so --only updates a subset in place.
    profiles: dict[str, dict] = {}
    if OUT.exists():
        profiles = json.loads(OUT.read_text())
    counts = {"ok": 0, "needs_browser": 0, "error": 0}

    for entry in entries:
        fid = entry["id"]
        rec = {"id": fid, "source_id": entry["source_id"], "name": entry["name"],
               "url": entry["url"], "kind": entry["kind"], "year": entry.get("year")}
        if entry.get("page"):
            rec["page"] = entry["page"]

        if needs_browser(entry["url"]):
            rec["status"] = "needs_browser"
            counts["needs_browser"] += 1
            profiles[fid] = rec
            print(f"  ~ {fid}: needs_browser (Box/Egnyte/Google)")
            continue

        path, err = (None, "skipped_no_download")
        if not args.no_download:
            path, err = download(entry, args.refresh)
        elif local_path(entry).exists():
            path, err = local_path(entry), None

        if err:
            rec["status"] = "error"
            rec["error"] = err
            counts["error"] += 1
            profiles[fid] = rec
            print(f"  ! {fid}: {err}")
            continue

        try:
            rec["profile"] = PROFILERS[entry["kind"]](path)
            rec["status"] = "ok"
            counts["ok"] += 1
            prof = rec["profile"]
            summary = (f"{prof['n_sheets']} tabs" if prof["kind"] == "xlsx"
                       else f"{prof.get('n_columns','?')} cols" if prof["kind"] == "csv"
                       else f"{prof['n_pages']} pages")
            print(f"  + {fid}: {summary}")
        except Exception as exc:  # noqa: BLE001
            rec["status"] = "error"
            rec["error"] = f"profile_failed: {exc}"
            counts["error"] += 1
            print(f"  ! {fid}: profile_failed: {exc}")

        profiles[fid] = rec

    OUT.write_text(json.dumps(profiles, indent=2, ensure_ascii=False))
    print(f"\nWrote {OUT.name}: {counts['ok']} ok, "
          f"{counts['needs_browser']} need browser, {counts['error']} errors.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
